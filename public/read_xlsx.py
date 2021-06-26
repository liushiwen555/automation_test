#!/bin/bash
# coding = utf-8
# Author: ZhaoGe
# Date: 2021/5/27 9:20

import openpyxl
import get_log
import pandas as pd

class ReadEmailXlsx(object):
    def __init__(self,filepath):
        self.book = openpyxl.load_workbook(filename)
        self.sheet = self.book.active
        self.rows_num = self.sheet.max_row
        self.cols_num = self.sheet.max_column
    def read_by_colums(self)->list:
        read_by_cols = []
        for col in self.sheet.iter_cols(max_col=self.cols_num):
            for cell in col:
                read_cell = cell.value
                read_by_cols.append(read_cell)
                print(read_by_cols)
        return read_by_cols












class ReadByOpenpyxl(object):
    def __init__(self,filename):
        self.book = openpyxl.load_workbook(filename)
        self.sheet = self.book.active
        self.rows_num = self.sheet.max_row
        self.cols_num = self.sheet.max_column
    def read_by_rows(self)->list:
        read_by_rows = []
        for row in self.sheet.iter_rows(max_row=self.rows_num):
            for cell in row:
                read_cell = cell.value
                read_by_rows.append(read_cell)
                print(read_cell, end='\t')
            print()
        return read_by_rows

    def read_by_colums(self)->list:
        read_by_cols = []
        for col in self.sheet.iter_cols(max_col=self.cols_num):
            for cell in col:
                read_cell = cell.value
                read_by_cols.append(read_cell)
                print(read_cell,end='\t')
            print()
        return read_by_cols
    def read_by_cells(self)->list:
        read_by_cells = []
        for i in range(self.rows_num):
            for j in range(self.cols_num):
                read_cell = self.sheet.cell(i+1,j+1).value
                read_by_cells.append(read_cell)
        return read_by_cells
if __name__ == '__main__':
    filename = r"F:\python_test\211code\seatworks\code\mail_info.xlsx"
    # load = ReadByOpenpyxl(filename)
    # print(load.read_by_rows())
    # print(load.read_by_colums())
    # print(load.read_by_cells())
    load = ReadEmailXlsx(filename)
    print(load.read_by_colums())